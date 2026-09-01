import re
import unicodedata
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import (
    DEFAULT_LEAVE_ALLOCATIONS,
    HalfDayPeriod,
    LeaveBalance,
    LeaveRequest,
    LeaveRequestDay,
    LeaveType,
    PublicHoliday,
    RequestStatus,
)

APPROVED_COLOR = '92D050'
PENDING_COLOR = 'F4B084'
HALF_DAY = Decimal('0.5')
LEGEND_MARKERS = ('jour féri', 'demande de cong', 'congé valid', 'conge valid')

# Excel spellings → canonical names as stored in the app (see User Management).
EXCEL_NAME_ALIASES = {
    'ABDERAHMAN ALOUNI': 'ABDERRAHIM ALOUINI',
    'SANA OUERGHEMI': 'SANA OUERGHEMMI',
}


def normalize_name(value: str) -> str:
    text = unicodedata.normalize('NFKD', value or '')
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return ' '.join(text.upper().split())


def cell_rgb(cell) -> str | None:
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = str(cell.fill.fgColor.rgb).upper()
            if len(rgb) >= 6:
                return rgb[-6:]
    except (TypeError, ValueError, AttributeError):
        return None
    return None


def leave_status_from_rgb(rgb: str | None) -> str | None:
    if rgb == APPROVED_COLOR:
        return RequestStatus.APPROVED
    if rgb == PENDING_COLOR:
        return RequestStatus.PENDING
    return None


def is_half_day(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)) == HALF_DAY
        except Exception:
            return False
    text = str(value).strip().lower()
    return '0,5' in text or '0.5' in text


def detect_year(sheet_title: str, file_path: Path, fallback: int) -> int:
    for source in (sheet_title, file_path.stem):
        match = re.search(r'(20\d{2})', source)
        if match:
            return int(match.group(1))
    return fallback


def build_column_dates(year: int, max_col: int) -> dict[int, date]:
    mapping: dict[int, date] = {}
    current = date(year, 1, 1)
    for col in range(2, max_col + 1):
        if current.year != year:
            break
        mapping[col] = current
        current += timedelta(days=1)
    return mapping


def is_employee_row(name: str) -> bool:
    lowered = name.lower()
    if not name or name == 'Nom':
        return False
    return not any(marker in lowered for marker in LEGEND_MARKERS)


def resolve_excel_name(name: str) -> str:
    normalized = normalize_name(name)
    return EXCEL_NAME_ALIASES.get(normalized, normalized)


def find_user_by_name(name: str) -> User | None:
    target = resolve_excel_name(name)
    candidates: list[tuple[User, str]] = []
    for user in User.objects.filter(is_active=True).select_related('profile'):
        full = normalize_name(user.get_full_name() or '')
        if full:
            candidates.append((user, full))

    for user, full in candidates:
        if full == target:
            return user

    target_parts = target.split()
    if len(target_parts) >= 2:
        swapped = f'{target_parts[-1]} {" ".join(target_parts[:-1])}'
        for user, full in candidates:
            if full == swapped:
                return user

    for user, full in candidates:
        full_parts = full.split()
        if len(full_parts) >= 2 and len(target_parts) >= 2:
            if full_parts[0] == target_parts[0] and full_parts[-1] == target_parts[-1]:
                return user

    return None


def get_or_create_balance(user, leave_type):
    balance, _ = LeaveBalance.objects.get_or_create(
        user=user,
        type=leave_type,
        defaults={
            'total': DEFAULT_LEAVE_ALLOCATIONS.get(leave_type, 0),
            'used': 0,
            'pending': 0,
        },
    )
    return balance


def request_level_period(entries):
    periods = {entry['half_day_period'] for entry in entries}
    if len(periods) == 1:
        return next(iter(periods))
    return None


def resolve_days(entries) -> Decimal:
    total = sum(
        HALF_DAY if entry['half_day_period'] else Decimal('1')
        for entry in entries
    )
    return total.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)


def group_entries_by_status(entries):
    grouped: dict[str, list[dict]] = {}
    for entry in entries:
        grouped.setdefault(entry['status'], []).append(entry)
    for status in grouped:
        grouped[status].sort(key=lambda item: item['date'])
    return grouped


class Command(BaseCommand):
    help = 'Importe les congés depuis un fichier Excel de planning annuel.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', help='Chemin vers le fichier .xlsx')
        parser.add_argument(
            '--year',
            type=int,
            help='Année du planning (défaut : détectée depuis le nom du fichier).',
        )
        parser.add_argument(
            '--leave-type',
            default=LeaveType.ANNUAL,
            choices=LeaveType.values,
            help='Type de congé à créer (défaut : annual).',
        )
        parser.add_argument(
            '--reason',
            default='Vacances',
            help='Raison enregistrée sur chaque demande importée.',
        )
        parser.add_argument(
            '--reviewer',
            help='E-mail de l’administrateur validateur (défaut : premier admin actif).',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Affiche le résumé sans écrire en base.',
        )
        parser.add_argument(
            '--skip-missing',
            action='store_true',
            help='Ignore les employés introuvables au lieu d’échouer.',
        )
        parser.add_argument(
            '--list-users',
            action='store_true',
            help='Affiche les employés actifs et quitte (utile pour vérifier les correspondances).',
        )

    def handle(self, *args, **options):
        if options['list_users']:
            self._print_active_users()
            return

        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError(
                'openpyxl est requis. Installez-le avec : pip install openpyxl'
            ) from exc

        excel_path = Path(options['excel_path']).expanduser()
        if not excel_path.is_file():
            raise CommandError(f'Fichier introuvable : {excel_path}')

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        year = options['year'] or detect_year(sheet.title, excel_path, date.today().year)
        column_dates = build_column_dates(year, sheet.max_column)
        holidays = set(
            PublicHoliday.objects.filter(date__year=year).values_list('date', flat=True)
        )

        reviewer = self._resolve_reviewer(options['reviewer'])
        leave_type = options['leave_type']
        reason = options['reason']
        dry_run = options['dry_run']

        parsed_rows = self._parse_sheet(sheet, column_dates, holidays)
        if not parsed_rows:
            raise CommandError('Aucun employé avec des congés trouvé dans le fichier.')

        created_requests = 0
        created_days = 0
        skipped_employees: list[str] = []
        employee_totals: list[tuple[str, Decimal, int]] = []

        with transaction.atomic():
            for employee_name, entries in parsed_rows:
                user = find_user_by_name(employee_name)
                if user is None:
                    if options['skip_missing']:
                        skipped_employees.append(employee_name)
                        self.stdout.write(
                            self.style.WARNING(f'Ignoré (introuvable) : {employee_name}')
                        )
                        continue
                    resolved = resolve_excel_name(employee_name)
                    raise CommandError(
                        f'Employé introuvable dans l’application : {employee_name} '
                        f'(recherché : {resolved}). '
                        'Créez le compte, vérifiez avec --list-users, ou utilisez --skip-missing.'
                    )

                matched_name = user.get_full_name() or user.username
                if normalize_name(matched_name) != resolve_excel_name(employee_name):
                    self.stdout.write(
                        f'Correspondance : {employee_name} -> {matched_name}'
                    )

                employee_counted = Decimal('0')
                employee_cells = 0

                for status, status_entries in group_entries_by_status(entries).items():
                    if self._has_overlap(user, status_entries):
                        raise CommandError(
                            f'Conflit détecté pour {employee_name} : '
                            'certaines dates existent déjà dans l’application.'
                        )

                    days = resolve_days(status_entries)
                    employee_counted += days
                    employee_cells += len(status_entries)
                    half_days = sum(
                        1 for entry in status_entries if entry['half_day_period']
                    )
                    date_labels = ', '.join(
                        f"{entry['date'].strftime('%d/%m')}"
                        f"{' (0,5J)' if entry['half_day_period'] else ''}"
                        for entry in status_entries
                    )

                    request = LeaveRequest(
                        employee=user,
                        type=leave_type,
                        start_date=status_entries[0]['date'],
                        end_date=status_entries[-1]['date'],
                        days=days,
                        half_day_period=request_level_period(status_entries),
                        status=status,
                        reason=reason,
                        emergency=False,
                    )
                    if status == RequestStatus.APPROVED:
                        request.reviewed_by = reviewer
                        request.reviewed_at = timezone.now()
                        request.review_comment = 'Import planning Excel'

                    if dry_run:
                        created_requests += 1
                        created_days += len(status_entries)
                        self.stdout.write(
                            f'[dry-run] {employee_name} -> {status} : '
                            f'{len(status_entries)} cellule(s), {days} jour(s) comptés'
                            + (f', {half_days} demi-journée(s)' if half_days else '')
                        )
                        self.stdout.write(f'           dates : {date_labels}')
                        continue

                    request.save()
                    LeaveRequestDay.objects.bulk_create(
                        [
                            LeaveRequestDay(
                                request=request,
                                date=entry['date'],
                                half_day_period=entry['half_day_period'],
                            )
                            for entry in status_entries
                        ]
                    )
                    balance = get_or_create_balance(user, leave_type)
                    if status == RequestStatus.APPROVED:
                        balance.used += days
                        balance.save(update_fields=['used'])
                    else:
                        balance.pending += days
                        balance.save(update_fields=['pending'])

                    created_requests += 1
                    created_days += len(status_entries)
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'{employee_name} -> {status} : '
                            f'{len(status_entries)} cellule(s), {days} jour(s) comptés'
                            + (f', {half_days} demi-journée(s)' if half_days else '')
                        )
                    )

                employee_totals.append((employee_name, employee_counted, employee_cells))

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write('')
        self.stdout.write('Récapitulatif par employé :')
        grand_total = Decimal('0')
        for employee_name, counted_days, cell_count in employee_totals:
            grand_total += counted_days
            self.stdout.write(
                f'  - {employee_name} : {counted_days} jour(s) comptés ({cell_count} cellule(s))'
            )
        self.stdout.write(f'  Total équipe : {grand_total} jour(s)')

        summary = (
            f'Import terminé : {created_requests} demande(s), '
            f'{created_days} jour(s) traité(s).'
        )
        if dry_run:
            summary = f'[dry-run] {summary} Aucune donnée enregistrée.'
        if skipped_employees:
            summary += f' Ignorés : {", ".join(skipped_employees)}.'
        self.stdout.write(self.style.SUCCESS(summary))

    def _print_active_users(self):
        self.stdout.write('Employés actifs dans l’application :')
        for user in User.objects.filter(is_active=True).order_by('first_name', 'last_name'):
            full = (user.get_full_name() or user.username).strip()
            self.stdout.write(f'  - {full} <{user.email}>')

    def _resolve_reviewer(self, reviewer_email: str | None) -> User | None:
        if reviewer_email:
            try:
                return User.objects.get(email__iexact=reviewer_email.strip(), is_active=True)
            except User.DoesNotExist as exc:
                raise CommandError(f'Reviewer introuvable : {reviewer_email}') from exc

        reviewer = (
            User.objects.filter(is_active=True, profile__role='admin')
            .order_by('id')
            .first()
        )
        if reviewer is None:
            reviewer = User.objects.filter(is_active=True, is_staff=True).order_by('id').first()
        return reviewer

    def _parse_sheet(self, sheet, column_dates, holidays):
        rows = []
        for row_idx in range(1, sheet.max_row + 1):
            raw_name = sheet.cell(row_idx, 1).value
            if raw_name is None:
                continue
            employee_name = str(raw_name).strip()
            if not is_employee_row(employee_name):
                continue

            entries = []
            for col, day in column_dates.items():
                if day.weekday() >= 5 or day in holidays:
                    continue
                rgb = cell_rgb(sheet.cell(row_idx, col))
                status = leave_status_from_rgb(rgb)
                if status is None:
                    continue

                half_day = is_half_day(sheet.cell(row_idx, col).value)
                entries.append(
                    {
                        'date': day,
                        'status': status,
                        'half_day_period': HalfDayPeriod.HALF if half_day else None,
                    }
                )

            if entries:
                rows.append((employee_name, entries))
        return rows

    def _has_overlap(self, user, entries):
        dates = [entry['date'] for entry in entries]
        return LeaveRequestDay.objects.filter(
            request__employee=user,
            date__in=dates,
        ).exclude(request__status=RequestStatus.REJECTED).exists()
